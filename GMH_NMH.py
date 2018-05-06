import openpyxl
import re
from nltk.corpus import stopwords

EXCEL_IN = 'forum.filtered.no.sent.length.xlsx'
EXCEL_OUT = 'mental.health.xlsx'


def filter_words(txt):
    if not txt:
        return False
    match = re.split(r'\s+', txt)
    common = ['scary', 'sleep', 'my head', 'things', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane', 'dog bark', 'social', 'delusions', 'sad','television', 'TV', 'hallucinations', 'time', 'meds', 'symptoms', 'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis', 'headaches']  # should be lowercase
    is_common = False
    for word in match:
        if word in common:
            is_common = True
            break
    return is_common


def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10):
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Non and General']
    sourceRow = 2
    bottomRow = len(posts['G']) + 1
    mentalRow = 2
    nonMentalRow = 2
    xl.create_sheet('General Mental Health')
    copy_to(1, 1, 'General Mental Health')  # copy headers
    xl.create_sheet('Non Mental Health')
    copy_to(1, 1, 'Non Mental Health')  # copy headers
    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value  # G column
        fw = filter_words(text)
        if fw:
            copy_to(sourceRow, mentalRow, 'General Mental Health')
            mentalRow += 1
        else:
            copy_to(sourceRow, nonMentalRow, 'Non Mental Health')
            nonMentalRow += 1
        sourceRow += 1
    xl.save(EXCEL_OUT)
    xl.close()
