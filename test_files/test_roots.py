import openpyxl
import re
from nltk.corpus import stopwords
import spacy
from nltk import Tree
from nltk.stem.snowball import SnowballStemmer
import nltk

EXCEL_IN = 'sample.xlsx'
EXCEL_OUT = 'test.xlsx'

obj = {
        'common': False,
        'personal': False,
        'long_count': False,
        'personal_check': False
    }

stop_words = stopwords.words('English')
common = ['scary', 'sleep', 'my head', 'things', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
          'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'time', 'meds', 'symptoms',
          'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis',
          'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed',
              'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'voices', 'sz', 'smell', 'daemon',
              'paranormal', 'pain', 'abilify', 'sedating', 'smell', 'demons', 'soul', 'nicotine', 'doctor', 'traumatized']  # should be lowercase

personal = ["i", "my", "i've", "i'm", "im", "ive", "me"]  # should be lowercase
stemmer = SnowballStemmer("english")
nlp = spacy.load('en')

def filter_object(text):

    if not text:
        return obj
    else:
        match = re.split(r'\s+', text)
    #check for start of sentence to identify personal comments
        if match[0].lower() in personal_roots:
            obj['personal'] = True
        count = 0

        for word in match:
            if word.lower() in common_roots: #lower the word to check against
                obj['common'] = True
            if "mg" in word:
                obj['common'] = True
            if len(word) > 2 and (word not in stop_words):
                count += 1
        if count > 2:
            obj['long_count'] = True

        return obj

def check_personal_comments(text):
    # final_list_stemmed = []
    # strip = re.split(r'\s+', text)
    # for w in strip:
    #     x = stemmer.stem(w)
    #     final_list_stemmed.append(x)
    # final_string_stemmed = ' '.join(final_list_stemmed)
    sent_text = nltk.sent_tokenize(text) #gives list of sentences
    # print(sent_text)
    print(sent_text)
    for se in sent_text:
        #print(type(se))

        doc = nlp((se))
        sentences = next(doc.sents)
        # print(type(sentences))
        for word in sentences:
            if word.dep_ == 'nsubj' and (str(word) in personal):
                obj['personal_check'] = True

    return obj


#simple copy function from one sheet to another row by row
def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10): #10 columns
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)

#only copies 7th column after applying filter method
def copy_to_v2(from_row, to_row, sheet_name):

    v = posts.cell(from_row, 7).value
    removed_words = filter(v)
    xl[sheet_name].cell(to_row, 7, removed_words)

if __name__ == '__main__':    # Script starts here

    print("starting categorization using roots")
    ################################################setup variables###########################################################
    # stop_words = stopwords.words('English')
    # common = ['scary', 'sleep', 'my head', 'things', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
    #       'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'time', 'meds', 'symptoms',
    #       'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis',
    #       'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed',
    #           'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'voices', 'sz', 'smell', 'daemon',
    #           'paranormal', 'pain', 'abilify', 'sedating', 'smell']  # should be lowercase
    #
    # personal = ["i", "my", "i've", "i'm", "im", "ive", "me"]  # should be lowercase
    # stemmer = SnowballStemmer("english")
    # nlp = spacy.load('en')

    common_roots = []
    personal_roots =[]


    for c in common:
        common_roots.append((stemmer.stem(c)))
    for p in personal:
       personal_roots.append(stemmer.stem(p))
    ################################################setup variables###########################################################

    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Cleaned_Posts']
    root_posts = xl['Stemmed_Posts']
    sourceRow = 3
    rootSourceRow = 3
    personalRow = 3
    commonRow = 3
    nonAndGenRow = 3
    gmhRow = 3
    nmhRow = 3

    bottomRow = len(root_posts['G']) + 1
    # xl.create_sheet('Personal')
    # copy_to(1, 1, 'Personal')  # copy headers
    xl.create_sheet('Personal & Common')
    copy_to(1, 1, 'Personal & Common')  # copy headers
    xl.create_sheet('Non and General')
    copy_to(1, 1, 'Non and General')  # copy headers
    # xl.create_sheet('General Mental Health')
    # copy_to(1, 1, 'General Mental Health')  # copy headers
    # xl.create_sheet('Non Mental Health')
    # copy_to(1, 1, 'Non Mental Health')  # copy headers
    while sourceRow < bottomRow:
        text = root_posts.cell(rootSourceRow, 7).value # G column from Cleaned_Posts Page
        text2 = posts.cell(sourceRow, 7).value
        fo = filter_object(text)
        fo = check_personal_comments(text2)
        # if ((fo['personal'] and fo['long_count']) or (fo['personal_check'] and fo['long_count'])):
        #     copy_to(sourceRow, personalRow, 'Personal')
        #     personalRow += 1
        #     print("Personal - processsed" + " " +str(sourceRow))
        # if fo['personal_check'] and fo['long_count']:
        #     copy_to(sourceRow, personalRow, 'Personal')
        #     personalRow += 1
        if ((fo['personal'] and fo['long_count'] and fo['common']) or (fo['personal_check'] and fo['long_count'] and fo['common'])):
            copy_to(sourceRow, commonRow, 'Personal & Common')
            commonRow += 1
            print("Personal/Common - processsed" + " " +str(sourceRow))

        else:
            copy_to(sourceRow, nonAndGenRow, 'Non and General')
            nonAndGenRow += 1
            print("Non -processsed" + " " +str(sourceRow))
        # if not fo['personal'] and fo['long_count'] and fo['common']:
        #     copy_to(sourceRow, gmhRow, 'General Mental Health')
        #     gmhRow += 1
        # if fo['personal'] or not fo['long_count'] or not fo['common']:
        #     copy_to(sourceRow, nmhRow, 'Non Mental Health')
        #     nmhRow += 1
        sourceRow += 1
        rootSourceRow += 1
        obj['common'] = False
        obj['long_count'] = False
        obj['personal'] = False
        obj['personal_check'] = False

    xl.save(EXCEL_OUT)
    xl.close()










