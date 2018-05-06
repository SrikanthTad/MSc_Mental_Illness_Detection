import openpyxl
from nltk.corpus import stopwords
from nltk.probability import FreqDist
import nltk

stop_words = stopwords.words ("english")

wb = openpyxl.load_workbook('sample.xlsx')
ws = wb.get_sheet_by_name('Personal')
mylist = []
legal_list=[]
for row in ws.iter_rows('G{}:G{}'.format(ws.min_row, ws.max_row)):
    for cell in row:
        # print ("Type: {}, Value: {}".format(type(cell.value), cell.value))
        # print(cell.value.split())
        for word in (cell.value.split()):
            word  = word.lower()
            if word not in stop_words:
                legal_list.append(word)

print(len(legal_list))

fdist = FreqDist(legal_list)
print(fdist.most_common(30))



#         for word in cell.value:
#             if word not in stop_words:
#                 legal_list.append(word)
# print(legal_list)
#         # mylist.append(cell.value)


