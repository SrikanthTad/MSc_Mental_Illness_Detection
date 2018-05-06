# import spacy
# from nltk import Tree
#
# nlp = spacy.load('en')
# doc = nlp("I've quick brown fox jumped over the moon")
# sentence = {}
# sentences = next(doc.sents)
# for word in sentences:
#     sentence[str(word)] = word.dep_
#     print(sentence[str(word)])
#     if sentence[str(word)] == "nsubj":
#         yes = True
# print(yes)
#
# print(sentence['fox']) #returns string value

import openpyxl
import re
from nltk.corpus import stopwords
import spacy
from nltk import Tree
from nltk.stem.snowball import SnowballStemmer

EXCEL_IN = 'forum.no.tags.xlsx'
EXCEL_OUT = 'forum.filtered.xlsx'

stop_words = stopwords.words('English')
common = ['scary', 'sleep', 'my head', 'things', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
          'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'time', 'meds', 'symptoms',
          'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis',
          'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed', 'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'voices', 'sz', 'smell', 'daemon']  # should be lowercase
personal = ["i", "my", "i've", "i'm", "im", "ive", "me"]  # should be lowercase


def filter_object(text):
    obj = {
        'common': False,
        'personal': False,
        'long_count': False,
        'personal_check': False
    }
    common_roots = []
    personal_roots =[]
    stemmer = SnowballStemmer("english")

    for c in common:
        common_roots.append((stemmer.stem(c)))
    for p in personal:
       personal_roots.append(stemmer.stem(p))
    if not text:
        return obj
    match = re.split(r'\s+', text)
    #check for start of sentence to identify personal comments
    if match[0].lower() in personal_roots:
        obj['personal'] = True
    count = 0
    #check for in sentence subjects to identify personal comments
    nlp = spacy.load('en')
    doc = nlp(str(text))
    sentence = {}
    sentences = next(doc.sents)
    for word in sentences:
        sentence[str(word)] = word.dep_
        if sentence[str(word)] == "nsubj" and (str(word) in personal):
            obj['personal_check'] = True


    for word in match:
        if word.lower() in common_roots: #lower the word to check against
            obj['common'] = True
        if "mg" in word:
            obj['common'] = True
        if len(word) > 2 and (word not in stop_words):
            count += 1
    if count > 4:
        obj['long_count'] = True
    return obj


def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10):
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)



if __name__ == '__main__':    # Script starts here
    print("starting categorization using roots")


    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']
    root_posts = xl['Cleaned_Posts']
    sourceRow = 3
    personalRow = 2
    commonRow = 2
    nonAndGenRow = 2
    gmhRow = 2
    nmhRow = 2
    rootSourceRow = 2
    bottomRow = len(root_posts['G']) + 1
    xl.create_sheet('Personal')
    copy_to(1, 1, 'Personal')  # copy headers
    xl.create_sheet('Personal & Common')
    copy_to(1, 1, 'Personal & Common')  # copy headers
    xl.create_sheet('Non and General')
    copy_to(1, 1, 'Non and General')  # copy headers
    xl.create_sheet('General Mental Health')
    copy_to(1, 1, 'General Mental Health')  # copy headers
    xl.create_sheet('Non Mental Health')
    copy_to(1, 1, 'Non Mental Health')  # copy headers
    while sourceRow < bottomRow:
        text = root_posts.cell(rootSourceRow, 7).value # G column from Cleaned_Posts Page
        fo = filter_object(text)
        if ((fo['personal'] and fo['long_count']) or (fo['personal_check'] and fo['long_count'])):
            copy_to(sourceRow, personalRow, 'Personal')
            personalRow += 1
            print("Personal - processsed" + " " +str(sourceRow))
        # if fo['personal_check'] and fo['long_count']:
        #     copy_to(sourceRow, personalRow, 'Personal')
        #     personalRow += 1
        if ((fo['personal'] and fo['long_count'] and fo['common']) or (fo['personal_check'] and fo['long_count'] and fo['common'])):
            copy_to(sourceRow, commonRow, 'Personal & Common')
            commonRow += 1
            print("Personal/Common - processsed" + " " +str(sourceRow))

        else:
            copy_to(sourceRow, nonAndGenRow, 'Non and General')
            nonAndGenRow += 1
            print("Non -processsed" + " " +str(sourceRow))
        # if not fo['personal'] and fo['long_count'] and fo['common']:
        #     copy_to(sourceRow, gmhRow, 'General Mental Health')
        #     gmhRow += 1
        # if fo['personal'] or not fo['long_count'] or not fo['common']:
        #     copy_to(sourceRow, nmhRow, 'Non Mental Health')
        #     nmhRow += 1
        sourceRow += 1
        rootSourceRow += 1
    xl.save(EXCEL_OUT)
    xl.close()










