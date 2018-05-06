import openpyxl
import re

EXCEL_IN = 'forum.schizophrenia.com.xlsx'
EXCEL_OUT = 'forum.no.tags.xlsx'


def remove_tags(html):
    match = re.split(r'(<[^>]+>)', html)
    result = []
    for tag in match:
        if len(result) % 2 == 0:
            result.append(tag)
        elif tag in ['<br>', '<hr>']:
            result.append('\n')
        elif tag.startswith('</'):
            result.append(' ')
        else:
            result.append('')
    return ''.join(result)


def add_protocol_prefix(links_short):
    links_full = []
    for link in links_short:
        matches = re.findall(r'www.youtube.com/embed/?([^?]+)', link)
        if len(matches) > 0:
            lf = "https://www.youtube.com/watch?v=" + matches[0]
        elif link.startswith('//forum.schizophrenia.com'):
            lf = 'https:' + link
        elif link.startswith('/letter_avatar/'):
            match = re.split(r'/', link)
            key = match[2].lower()
            if key in users:
                match[2] = str(users[key])
            lf = '/'.join(match)
        elif '/user_avatar/forum.schizophrenia.com/' in link:
            match = re.split(r'/', link)
            u = match.index('user_avatar') + 2
            key = match[u].lower()
            if key in users:
                match[u] = str(users[key])
            lf = '/'.join(match)
        elif link.startswith('/users/'):
            key = link[7:].lower()
            if key in users:
                key = str(users[key])
            lf = '/users/' + key
        elif link.startswith('/u/'):
            key = link[3:].lower()
            if key in users:
                key = str(users[key])
            lf = '/u/' + key
        else:
            lf = link
        links_full.append(lf)
    return links_full


def add_embeds(src):
    links = []
    ids = re.findall(r'data-youtube-id=[\'"]?([^\'" >]+)', src)
    for i in ids:
        links.append("https://www.youtube.com/watch?v=" + i)
    return links


def substitute_user_ids(src):
    match = re.split(r'(@\w+)', src)
    result = []
    for w in match:
        if len(result) % 2 == 0:
            result.append(w)
        else:
            key = w[1:].lower()
            if key in users:
                result.append('@' + str(users[key]))
            else:
                result.append(key)
    return ''.join(result)


def index_of(posts):
    ids = {}
    for r in range(2, len(posts['E'])):
        key = posts.cell(r, 6).value
        if not key:
            continue
        key = key.lower()
        if not key in ids:
            ids[key] = posts.cell(r, 5).value
    return ids


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    users = index_of(xl['Posts'])
    empty_row = len(xl['Posts']['C']) + 1
    xl['Posts'].cell(1, 7, 'strip')
    xl['Posts'].cell(1, 8, 'hrefs')
    xl['Posts'].cell(1, 9, 'image URLs')
    for r in range(2, empty_row):
        src = xl['Posts'].cell(r, 3).value
        if src is not None:
            strip = remove_tags(src)
            # Sanitize bad leading characters
            if len(strip) > 0:
                a = ord(strip[0])
                if a > 250 or a == 61:
                    strip = ' ' + strip
                strip = substitute_user_ids(strip)
            xl['Posts'].cell(r, 7, strip)
            hrefs = re.findall(r'href=[\'"]?([^\'" >]+)', src)
            hrefs = add_protocol_prefix(hrefs) + add_embeds(src)
            xl['Posts'].cell(r, 8, '\n\n'.join(hrefs))
            images = re.findall(r'src=[\'"]?([^\'" >]+)', src)
            images = add_protocol_prefix(images)
            xl['Posts'].cell(r, 9, '\n\n'.join(images))
        if r % 200 == 0:
            print('Progress: {}'.format(r))
    xl.save(EXCEL_OUT)
    xl.close()
