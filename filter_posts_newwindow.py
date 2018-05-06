import openpyxl
import re
from nltk.corpus import stopwords

EXCEL_IN = 'forum.no.tags.xlsx'
EXCEL_OUT = 'forum.filtered.xlsx'
EXCEL_PS = 'personal.xlsx'


def first_word(text):
    if not text:
        return False
    match = re.split(r'\s+', text)
    count = 0
    stop_words = stopwords.words('english')
    for word in match:
        if len(word) > 2 and (word not in stop_words):
            count += 1
            if count > 5:
                # this my be 'i' or 'my' or whichever
                return match[0]
    return False


def copy(src, tgt):
    for i in range(1, 10):
        v = posts.cell(src, i).value
        posts.cell(tgt, i, v)
        posts.cell(src, i, '')


def clear(src):
    for i in range(1, 10):
        posts.cell(src, i, '')


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']
    targetRow = 2
    sourceRow = 3
    bottomRow = len(posts['G']) + 1
    personal = []
    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value # G column
        fw = first_word(text)
        if fw:
            copy(sourceRow, targetRow)
            if fw.lower() in ['i', 'my']:
                personal.append(targetRow)
            targetRow += 1
        else:
            clear(sourceRow)
            print('"{}"'.format(text))
        sourceRow += 1
    xl.save(EXCEL_OUT)
    # print(personal)
    targetRow = 2
    sourceRow = 2
    bottomRow = len(posts['G']) + 1
    while sourceRow < bottomRow:
        if sourceRow in personal:
            copy(sourceRow, targetRow)
            targetRow += 1
        else:
            clear(sourceRow)
        sourceRow += 1
    xl.save(EXCEL_PS)
    xl.close() #TODO: make process faster lololol

# pronouns and specific words
# tag and put specific rows with colors
# use Hanan's words
