import spacy
from nltk.corpus import stopwords
import openpyxl
import re
from nltk.corpus import stopwords
import spacy
from nltk import Tree
from nltk.stem.snowball import SnowballStemmer
import string

EXCEL_IN = 'sample.xlsx'
EXCEL_OUT = 'sample.xlsx'


common = ['scary', 'sleep', 'head', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
          'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'meds', 'symptoms',
          'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis', 'psychotic',
          'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed', 'counselor', 'pychologist',
          'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'schizophrenic', 'voices', 'sz', 'daemon', 'fear',
          'paranormal', 'pain', 'abilify', 'sedating', 'smell', 'demons', 'soul', 'nicotine', 'doctor', 'traumatized',
          'telephone', 'ringing', 'agitation', 'light', 'stress', 'derealization', 'god', 'pdoc', 'dream', 'ra', 'problem', 'episodes',
          'feel', 'danger', 'insomnia', 'anxiety', 'disorder', 'sza', 'smoke', 'weight', 'antidepressants', 'adhd', 'concerta', 'clozapine',
          'dose', 'xanax', 'levitate', 'wonder', 'syndrome', 'ward', 'pill', 'depression', 'zopiclone', 'paranoia', 'normies', 'mood', 'depression',
          'struggle', 'brain', 'daydream', 'schizoaffective', 'invega', 'provigil', 'geoden', 'seroquel', 'chlorpromazine', 'depixol', 'latuda', 'loxitane',
          'trifluperazine', 'prolixin', 'haloperidol', 'clonidine', 'mentally']  # should be lowercase

#stems and removes useless words from stemmed posts page
def filter(text):
    remove_list = ["szadmin", "szadmin's", "(szadmin)", "radmatech", "szadmin.", "szadmin:", "szadmin,", "sz admin",
                     "surprised j", "surprisedJ", 'surprisedj:', "surprisedj said:", " RowanAmethyst:", "RowanAmethyst", "szadmin,"]
    final_list = []
    s = 'null'
    #final_string_stemmed =[]
    final_list_stemmed = []

    if not text:
        return None
    else:
        tt = re.sub('['+string.punctuation+']', '', text)
        #tt = re.sub('(?<=\D)[.,]|[.,](?=\D)', '', text)
        new_text = re.split(r'\s+', tt)
        for word in new_text:
            if (word.lower()) not in remove_list:
                x = stemmer.stem(word)
                final_list_stemmed.append(x)
        final_string_stemmed = ' '.join(final_list_stemmed)
        print(final_string_stemmed)
        return final_string_stemmed

#removes useless words only from posts page
def remove_useless_words(text):
    useless_words = ["szadmin", "szadmin's", "(szadmin)", "radmatech", "szadmin.", "szadmin:", "szadmin,", "sz admin",
                     "surprised j", "surprisedJ", 'surprisedj:', "surprisedj said:", " RowanAmethyst:", "RowanAmethyst"]
    final_lists = []
    if not text:
        return None
    else:
        stripped_text = re.split(r'\s+', text)
        for word in stripped_text:
            if word.lower() not in useless_words:
                final_lists.append(word)
        final_string = ' '.join(final_lists)
        return final_string

#simple copy function from one sheet to another row by row
def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10):
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)

#only copies 7th column after applying filter method
def copy_to_v2(from_row, to_row, sheet_name):

    v = posts.cell(from_row, 7).value
    removed_words = filter(v)
    xl[sheet_name].cell(to_row, 7, removed_words)

#only copies 7th column after applying remove_useless_words method
def copy_to_v3(from_row, to_row, sheet_name):
    v = posts.cell(from_row,7).value
    fin = remove_useless_words(v)
    xl[sheet_name].cell(to_row, 7, fin)

#main
if __name__ == "__main__":
    print("removing undesired words and stemming...")


    stemmer = SnowballStemmer("english")


    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']

    sourceRow = 3 #workbook data starts at 3 - start of file
    removedRow = 3
    stemmedRow = 3
    bottomRow = len(posts['G']) + 1 #end of file

    xl.create_sheet("Stemmed_Posts")
    copy_to(1,1,'Stemmed_Posts') #copy headers
    xl.create_sheet("Cleaned_Posts")
    copy_to(1,1,'Cleaned_Posts') #copy headers

    while sourceRow < bottomRow:
       # text = posts.cell(sourceRow, 7).value # G column
        copy_to(sourceRow, removedRow, 'Cleaned_Posts') #copy them regularly first so we can overwrite the 7th column only after removing useless words
        copy_to_v2(sourceRow, stemmedRow, 'Stemmed_Posts')
        copy_to_v3(sourceRow, removedRow, 'Cleaned_Posts')
        removedRow +=1
        sourceRow +=1
        stemmedRow +=1

    xl.save(EXCEL_OUT)
    xl.close()







