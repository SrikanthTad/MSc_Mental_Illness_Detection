import spacy
from nltk.corpus import stopwords
import openpyxl
import re
from nltk.corpus import stopwords
import spacy
from nltk import Tree
from nltk.stem.snowball import SnowballStemmer

EXCEL_IN = 'forum.no.tags.xlsx'
EXCEL_OUT = 'forum.no.tags.xlsx'

# trigger = False
# nlp = spacy.load('en')
# doc = nlp("I would eat malmalade but I wouldn't eat raw orange peel")
# sentence = {}
# sentences = next(doc.sents)
# for word in sentences:
#     sentence[str(word)] = word.dep_
#     print(sentence[str(word)])
#     if sentence[str(word)] == "nsubj":
#         trigger = True
#         # obj['personal_check'] = True


# print(trigger)

common = ['scary', 'sleep', 'my head', 'things', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
          'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'time', 'meds', 'symptoms',
          'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis',
          'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed', 'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'voices', 'sz']


def filter(text):
    remove_list = ["szadmin", "szadmin's", "(szadmin)", "radmatech", "szadmin.", "szadmin:"]
    final_list = []
    s = 'null'
    #final_string_stemmed =[]
    final_list_stemmed = []

    if not text:
        return s
    if text is not None:
        new_text = re.split(r'\s+', text)
        for word in new_text:
            if (word.lower()) not in remove_list:
            # print(word)
                final_list.append(word)
        for x in final_list:
            x2 = stemmer.stem(x)
            final_list_stemmed.append(x2)
        final_string_stemmed = ' '.join(final_list_stemmed)
        print(final_string_stemmed)

        return final_string_stemmed

# print(common_stemmed)
#
# def filter_object(text):
#     remove_list = ["SzAdmin"]
#     obj = {
#         'common': False,
#         'personal': False,
#         'long_count': False,
#         'personal_check': False
#     }
#     if not text:
#         return obj
#     match = re.split(r'\s+', text)
#     #check for start of sentence to identify personal comments
#     for word in match:
#         if
#
#
#
def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10):
        #posts.cell(from_row, 7).value = filter(posts.cell(from_row, 7).value)

        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)

def copy_to_v2(from_row, to_row, sheet_name):

        #posts.cell(from_row, 7).value = filter(posts.cell(from_row, 7).value)

        v = posts.cell(from_row, 7).value
        removed_words = filter(v)

        xl[sheet_name].cell(to_row, 7, removed_words)

if __name__ == "__main__":
    print("removing undesired words")

    common_stemmed = []


    stemmer = SnowballStemmer("english")
    for w in common:
        common_stemmed.append(stemmer.stem(w))

    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']

    sourceRow = 3 #workbook data starts at 3 - start of file
    removedRow = 2
    bottomRow = len(posts['G']) + 1 #end of file

    xl.create_sheet("Cleaned_Posts")
    copy_to(1,1,'Cleaned_Posts') #copy headers

    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value # G column
        #text = filter(text)
        copy_to(sourceRow, removedRow, 'Cleaned_Posts')
        copy_to_v2(sourceRow, removedRow, 'Cleaned_Posts')
        removedRow +=1
        sourceRow +=1

    xl.save(EXCEL_OUT)
    xl.close()






#     while sourceRow < bottomRow:
#         text = posts.cell(sourceRow, 7).value # G column
#         fo = filter_object(text)
#         if ((fo['personal'] and fo['long_count']) or (fo['personal_check'] and fo['long_count'])):
#             copy_to(sourceRow, personalRow, 'Personal')
#             personalRow += 1
#             print("Personal - processsed" + " " +str(sourceRow))
#         # if fo['personal_check'] and fo['long_count']:
#         #     copy_to(sourceRow, personalRow, 'Personal')
#         #     personalRow += 1
#         if ((fo['personal'] and fo['long_count'] and fo['common']) or (fo['personal_check'] and fo['long_count'] and fo['common'])):
#             copy_to(sourceRow, commonRow, 'Personal & Common')
#             commonRow += 1
#             print("Personal/Common - processsed" + " " +str(sourceRow))
#
#         else:
#             copy_to(sourceRow, nonAndGenRow, 'Non and General')
#             nonAndGenRow += 1
#             print("Non -processsed" + " " +str(sourceRow))
#         # if not fo['personal'] and fo['long_count'] and fo['common']:
#         #     copy_to(sourceRow, gmhRow, 'General Mental Health')
#         #     gmhRow += 1
#         # if fo['personal'] or not fo['long_count'] or not fo['common']:
#         #     copy_to(sourceRow, nmhRow, 'Non Mental Health')
#         #     nmhRow += 1
#         sourceRow += 1
#     xl.save(EXCEL_OUT)
#     xl.close()
#
#
#
#
#
#
#
#
#
#
