import openpyxl
import requests
import json
import time
from datetime import datetime

# Project constants: URL templates and local file name in the script directory
TOPICS_URL = 'https://forum.schizophrenia.com/latest.json?page={}'
POSTS_URL = 'https://forum.schizophrenia.com/t/{}.json'
EXCEL_FILE = 'forum.schizophrenia.com.xlsx'

# We start scraping from the most old pages available.
# We also exclude pages, scraped earlier.
# So we must find the last page P for start scraping and then
# do sequencial requests with parameter P, P-1, P-2, ...
# This function return this page number - P, - beginning with p value (in argument)
def from_page(p):
    delta = 15  # long step (quick) for search iterations
    if p < 15:
        delta = 1
    while p > -1:       # Search iterations loop
        re = requests.get(TOPICS_URL.format(p))
        if not re.status_code == 200:
            return 0    # Abnormal server answer
        j = json.loads(re.text)
        if len(j['topic_list']['topics']) == 0:
            print('{} server has no such page'.format(p))
            p -= delta  # Page number appears too big
                        # and server returns no topics data
        else:
            # Server answer contains topics data and we must look at topic dates
            page_datetime = utc_datetime(j['topic_list']['topics'][0])
            # All topics with date older than settings['last_datetime']
            # are already treated and not interesting for us
            if page_datetime < settings['last_datetime']:
                print('{} already treated'.format(p))
                settings['start_page'] = p
                p -= delta     # Search decrement
            elif delta == 1:
                print('{} is chosen as start page'.format(p))
                return p
            else:
                # This page must be treated, but it is found by long step
                # We must reconsider close pages to avoid they untreated
                p += delta - 1
                delta = 1   # shortest iteration step (slow)
    return 0    # this case is for fully scraped forum


# Convert appropriate datetime from JSON UTC to Py
def utc_datetime(topic):
    if topic['last_posted_at'] is None:
        d = topic['created_at']
    else:
        d = topic['last_posted_at']
    return datetime.strptime(d, '%Y-%m-%dT%H:%M:%S.%fZ')


# Topic data extraction
def scrape_topic(topic):
    # Is the topic among scraped data?
    if topic['id'] in topics_index:
        topic_row = topics_index.index(topic['id']) + 2
    else:
        topics_index.append(topic['id'])
        topic_row = len(topics_index) + 2   # blank row in excel sheet
    topic_datetime = utc_datetime(topic)
    # We modify settings for future work, avoiding double data treatment
    if topic_datetime > settings['last_datetime']:
        settings['last_datetime'] = topic_datetime
    # Excel row is updated with topic data
    xl['Topics'].cell(topic_row, 1, topic['category_id'])
    xl['Topics'].cell(topic_row, 2, topic['id'])
    xl['Topics'].cell(topic_row, 3, topic['title'])
    xl['Topics'].cell(topic_row, 4, topic_datetime)


# Post data extraction
def scrape_post(post):
    # Is the post already scraped?
    # I think: always NO. This part of code may be surplus...
    if post['id'] in posts_index:
        post_row = posts_index.index(post['id']) + 2
    else:
        posts_index.append(post['id'])
        post_row = len(posts_index) + 2 # blank row for new post data
    post_datetime = datetime.strptime(post['updated_at'], '%Y-%m-%dT%H:%M:%S.%fZ')
    # Excel row is updated with post data
    xl['Posts'].cell(post_row, 1, post['topic_id'])
    xl['Posts'].cell(post_row, 2, post['id'])
    xl['Posts'].cell(post_row, 3, post['cooked'])
    xl['Posts'].cell(post_row, 4, post_datetime)
    xl['Posts'].cell(post_row, 5, post['user_id'])
    xl['Posts'].cell(post_row, 6, post['username'])


# We create previously scraped objects index
def index_of(sheet_name):
    ids = []
    for r in range(2, len(xl[sheet_name]['B'])):
        ids.append(xl[sheet_name].cell(r, 2).value)
    return ids


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_FILE)
    # Read setting data from Excel
    settings = {
        'time_limit': xl['Settings'].cell(2, 2).value,
        'last_datetime': xl['Settings'].cell(3, 2).value,
        'start_page': xl['Settings'].cell(4, 2).value
    }
    # Prepare indices of already scraped objects
    topics_index = index_of('Topics')
    posts_index = index_of('Posts')

    # Find starting page URL parameter
    page = from_page(settings['start_page'])
    # Setup time limit for script work
    end_time = time.time() + settings['time_limit'] * 60
    while time.time() < end_time:
        sec = int(end_time - time.time())
        # Current time info and active page in progress
        print('{} sec rest\n{} in work'.format(sec, page))
        response = requests.get(TOPICS_URL.format(page))
        if response.status_code == 200:
            topics_json = json.loads(response.text)
            # Append topics data from JSON server answer
            for topic in topics_json['topic_list']['topics']:
                print('\tTopic {} from page {}'.format(topic['id'], page))
                scrape_topic(topic)
                re = requests.get(POSTS_URL.format(topic['id']))
                if re.status_code == 200:
                    posts_json = json.loads(re.text)
                    # Append posts data from JSON server answer
                    for post in posts_json['post_stream']['posts']:
                        # print('\t\tPost {}'.format(post['id']))
                        scrape_post(post)
        page -= 1   # Page decrement for next request
        if page < 0:
            break

    # Save results in file after work
    xl['Settings'].cell(3, 2, settings['last_datetime'])
    xl.save(EXCEL_FILE)
    xl.close()
