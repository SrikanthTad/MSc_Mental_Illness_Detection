import openpyxl
import re
from nltk.corpus import stopwords

EXCEL_IN = 'forum.no.tags.xlsx'
EXCEL_OUT = 'forum.filtered.xlsx'


def filter_words(text):
    if not text:
        return False
    match = re.split(r'\s+', text)
    count = 0
    stop_words = stopwords.words('english')
    personal = ['i', 'my'] # should be lowercase
    common = ['television', 'TV' 'hallucinations', 'time', 'meds', 'symptoms', 'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared'] # should be lowercase
    is_common = False
    for word in match:
        if len(word) > 2 and (word not in stop_words):
            count += 1
        if word in common:
            is_common = True
    if count < 6:
        return False
    # this my be 'i' or 'my' or something else
    return [match[0].lower() in personal, is_common]


def copy():
    for i in range(1, 10):
        v = posts.cell(sourceRow, i).value
        posts.cell(targetRow, i, v)
        posts.cell(sourceRow, i, '')


def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10):
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)


def clear(src):
    for i in range(1, 10):
        posts.cell(src, i, '')


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']
    targetRow = 2
    sourceRow = 3
    personalRow = 2
    commonRow = 2
    bottomRow = len(posts['G']) + 1
    xl.create_sheet('Personal')
    copy_to(1, 1, 'Personal')  # copy headers
    xl.create_sheet('Personal & Common')
    copy_to(1, 1, 'Personal & Common')  # copy headers
    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value # G column
        fw = filter_words(text)
        if fw:
            copy()
            if fw[0]:
                copy_to(targetRow, personalRow, 'Personal')
                personalRow += 1
                if fw[1]:
                    copy_to(targetRow, commonRow, 'Personal & Common')
                    commonRow += 1
            targetRow += 1
        else:
            clear(sourceRow)
            print('"{}"'.format(text))
        sourceRow += 1
    xl.save(EXCEL_OUT)
    xl.close()
