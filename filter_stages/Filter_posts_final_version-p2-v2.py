import openpyxl
import re
from nltk.corpus import stopwords
import spacy
from nltk import Tree
from nltk.stem.snowball import SnowballStemmer
import nltk

EXCEL_IN = 'sample.xlsx'
EXCEL_OUT = 'sample.xlsx'

obj = {
        'common': False
    }

stop_words = stopwords.words('English')
common = ['scary', 'sleep','hippocampus', 'head', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
          'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'meds', 'symptoms',
          'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis', 'psychotic',
          'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed', 'counselor', 'pychologist',
          'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'schizophrenic', 'voices', 'sz', 'daemon', 'fear',
          'paranormal', 'pain', 'abilify', 'sedating', 'smell', 'demons', 'soul', 'nicotine', 'doctor', 'traumatized',
          'telephone', 'ringing', 'agitation', 'light', 'stress', 'derealization', 'god', 'pdoc', 'dream', 'ra', 'problem', 'episodes',
          'feel', 'danger', 'insomnia', 'anxiety', 'disorder', 'sza', 'smoke', 'weight', 'antidepressants', 'adhd', 'concerta', 'clozapine',
          'dose', 'xanax', 'levitate', 'wonder', 'syndrome', 'ward', 'pill', 'depression', 'zopiclone', 'paranoia', 'normies', 'mood', 'depression',
          'struggle', 'brain', 'daydream', 'schizoaffective', 'invega', 'provigil', 'geoden', 'seroquel', 'chlorpromazine', 'depixol', 'latuda', 'loxitane',
          'trifluperazine', 'prolixin', 'haloperidol', 'clonidine', 'mentally', 'stigma', 'schizos', 'paliperidone', 'selfmedicate', 'tomography']  # should be lowercase
common_bigrams = ['mentally ill', 'severe mental', 'my mental', 'the paranoia', 'health care', 'heard voices', 'for anxiety', 'life was', 'depressed and', 'medication for'
                  'your psychiatrist', 'my meds', 'side effects', 'the meds', 'meds I', 'with schizophrenia', 'on meds', 'my therapist', 'paliperidone palmitate', 'psychosis FEP',
                  'brain nicotine', 'battery mccb', 'smoking schizophrenia', 'cerebral diabetics', 'diabetic brain', 'by antipsychotic', '20 mg', 'schizophrenia sz',
                  'cognitive remediation', 'paranoid schizophrenia', 'bipolar depression', 'auditory hallucinations', 'negative psychotic', 'increased pain', 'medication refractory'
                  , 'anger control', 'sleep patters', 'anxiety disorder', 'cbd meds', 'distress tolerance', 'symptom remission', 'prolactin levels', 'of risperidone',
                  'schizophrenia you', 'atypical antipsychotics', 'hippocampus and', 'lose weight', 'antipsychotic agents', 'first episode', 'firstepisode', 'selfmedicate symptoms',
                  'brain waves', 'positron emission', 'drug that', 'to antipsychotics', 'psychotic episode', 'dosages of', 'including schizophrenia', 'treatment approaches',
                  'psychiatric symptoms', 'spectrum disorders', 'severe mental', '95 ci', 'of aripiprazole' 'of clozapine']

personal = ["i", "my", "i've", "i'm", "im", "ive", "me", "mine"]  # should be lowercase
stemmer = SnowballStemmer("english")
nlp = spacy.load('en')

def filter_object(text):

    if not text:
        return obj
    else:
        # match = re.sub('(?<=\D)[.,]|[.,](?=\D)', '', text)
        match = re.split(r'\s+', text)
        #print(match)
    #check for start of sentence to identify personal comments
        # if match[0].lower() in personal_roots:
        #     obj['personal'] = True
        #count = 0
        for v in common_roots_bigrams:
            if v.lower() in text.lower():
                obj['common'] = True
        for word in match:
            if word.lower() in common_roots: #lower the word to check against
                obj['common'] = True
            if ("mg" or "thought" or 'dreams') in word:
                obj['common'] = True
        #     if len(word) > 2 and (word not in stop_words):
        #         count += 1
        # if count > 2:
        #     obj['long_count'] = True
        print(obj)
        return obj

# def check_personal_comments(text):
#     # final_list_stemmed = []
#     # strip = re.split(r'\s+', text)
#     # for w in strip:
#     #     x = stemmer.stem(w)
#     #     final_list_stemmed.append(x)
#     # final_string_stemmed = ' '.join(final_list_stemmed)
#     sent_text = nltk.sent_tokenize(text) #gives list of sentences
#     # print(sent_text)
#     for se in sent_text:
#         #print(type(se))
#         # don't need to apply special character removal since it's handled by spacy interpreter
#         doc = nlp((se))
#         sentences = next(doc.sents)
#         # print(type(sentences))
#         for word in sentences:
#             if word.dep_ == 'nsubj' and (str(word).lower() in personal):
#                 obj['personal_check'] = True
#
#     return obj


#simple copy function from one sheet to another row by row
def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10): #10 columns
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)

#only copies 7th column after applying filter method
# def copy_to_v2(from_row, to_row, sheet_name):
#
#     v = posts.cell(from_row, 7).value
#     removed_words = filter(v)
#     xl[sheet_name].cell(to_row, 7, removed_words)

if __name__ == '__main__':    # Script starts here

    print("starting categorization using roots")

    ################################################setup variables###########################################################
    # stop_words = stopwords.words('English')
    # common = ['scary', 'sleep', 'my head', 'things', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
    #       'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'time', 'meds', 'symptoms',
    #       'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis',
    #       'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed',
    #           'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'voices', 'sz', 'smell', 'daemon',
    #           'paranormal', 'pain', 'abilify', 'sedating', 'smell']  # should be lowercase
    #
    # personal = ["i", "my", "i've", "i'm", "im", "ive", "me"]  # should be lowercase
    # stemmer = SnowballStemmer("english")
    # nlp = spacy.load('en')

    # common_roots = []
    # personal_roots =[]
    #
    #
    # for c in common:
    #     common_roots.append((stemmer.stem(c)))
    # for p in personal:
    #    personal_roots.append(stemmer.stem(p))
    ################################################setup variables###########################################################
    common_roots = []
    personal_roots =[]
    common_roots_bigrams = []

    for c in common:
        common_roots.append((stemmer.stem(c)))
    for p in personal:
        personal_roots.append(stemmer.stem(p))
    for k in common_bigrams:
        common_roots_bigrams.append((stemmer.stem(k)))

    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Non and General']
    root_posts = xl['Stemmed_Non and General']
    sourceRow = 3
    rootSourceRow = 3
    personalRow = 3
    commonRow = 3
    nonAndGenRow = 3
    gmhRow = 3
    nmhRow = 3

    bottomRow = len(root_posts['G']) + 1
    # xl.create_sheet('Personal')
    # copy_to(1, 1, 'Personal')  # copy headers
    # xl.create_sheet('General Mental Health')
    # copy_to(1, 1, 'General Mental Health')  # copy headers
    # xl.create_sheet('Non Mental Health')
    # copy_to(1, 1, 'Non Mental Health')  # copy headers
    xl.create_sheet('General Mental Health')
    copy_to(1, 1, 'General Mental Health')  # copy headers
    xl.create_sheet('Non Mental Health')
    copy_to(1, 1, 'Non Mental Health')  # copy headers
    while sourceRow < bottomRow:
        text = root_posts.cell(rootSourceRow, 7).value # G column from Cleaned_Posts Page

        #text2 = posts.cell(sourceRow, 7).value
        fo = filter_object(text)

        # fo = check_personal_comments(text2)
        # if ((fo['personal'] and fo['long_count']) or (fo['personal_check'] and fo['long_count'])):
        #     copy_to(sourceRow, personalRow, 'Personal')
        #     personalRow += 1
        #     print("Personal - processsed" + " " +str(sourceRow))
        # if fo['personal_check'] and fo['long_count']:
        #     copy_to(sourceRow, personalRow, 'Personal')
        #     personalRow += 1
        if fo['common']:
            copy_to(sourceRow, gmhRow, 'General Mental Health')
            gmhRow += 1
            print("General mental health - processsed" + " " +str(sourceRow))

        else:
            copy_to(sourceRow, nmhRow, 'Non Mental Health')
            nmhRow += 1
            print("Non mental health -processsed" + " " +str(sourceRow))

        # if not fo['personal'] and fo['long_count'] and fo['common']:
        #     copy_to(sourceRow, gmhRow, 'General Mental Health')
        #     gmhRow += 1
        # if fo['personal'] or not fo['long_count'] or not fo['common']:
        #     copy_to(sourceRow, nmhRow, 'Non Mental Health')
        #     nmhRow += 1
        sourceRow += 1
        rootSourceRow += 1

        obj['common'] = False



    xl.save(EXCEL_OUT)
    xl.close()










