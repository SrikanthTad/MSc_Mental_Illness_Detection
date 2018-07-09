import openpyxl
import re
from nltk.util import ngrams
import string

from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords

from nltk.probability import FreqDist
import nltk

EXCEL_IN = 'sample.xlsx'
# EXCEL_IN = 'forum.filtered.xlsx'
#EXCEL_OUT = 'forum.filtered.xlsx'
slist = []
tlist=[]
flist =[]
legal_list =[]

stop_words = stopwords.words("english")

def filter_words(text):
    # if not text:
    #     return False
    match = re.split(r'\s+', text)
    return match





if __name__ == '__main__':    # Script starts here
    legal_list = []
    new_list = []
    wb = openpyxl.load_workbook('sample.xlsx')
    sourceRow = 3
    rootSourceRow = 3
    personalRow = 3
    commonRow = 3
    nonAndGenRow = 3
    gmhRow = 3
    nmhRow = 3

    posts = wb['General Mental Health']
    bottomRow = 40000
        # len(posts['G']) + 1

    while sourceRow < bottomRow:
            text = posts.cell(sourceRow, 7).value # G column from Cleaned_Posts Page
            # print(type(text))
            if(text == None):
                pass
            else:

    #
    # wb = openpyxl.load_workbook('sample.xlsx')
    # ws = wb.get_sheet_by_name('Posts')
    # mylist = []
    # for row in ws.iter_rows('G{}:G{}'.format(ws.min_row, ws.max_row)):
    #     for cell in row:
    #         if (cell.value == None):
    #             pass
    #         else:
                # print(cell.value)
            # print ("Type: {}, Value: {}".format(type(cell.value), cell.value))
            #     for sent in tokenize
            #     tokenize = nltk.word_tokenize(text)
            #     print(tokenize)
            #     for w in tokenize:
            #         legal_list.append(w)
                text_stripped = re.sub('['+string.punctuation+']', '', text)
                tokenize = nltk.sent_tokenize(text_stripped)

                for sent in tokenize:
                    tokenize_words = nltk.word_tokenize((sent))
                    for w in tokenize_words:

                        new_list.append(w)

            sourceRow += 1

    print("============================================SPLIT===============================================")


    bgs = (ngrams(new_list,3))
    fdist = nltk.FreqDist(bgs)
    for l in (fdist.most_common(5000)):
        print(l)
    # for k,v in fdist.items():
    #     print (k,v)
            #     for word in (cell.value.split()):
            #         word = word.lower()
            #         legal_list.append(word)
                # print(legal_list)


print("============================================SPLIT===============================================")


# fdist = FreqDist(legal_list)
# #print(fdist.most_common(200))
# for l in (fdist.most_common(200)):
#     print(l)
