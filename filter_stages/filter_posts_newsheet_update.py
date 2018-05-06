import openpyxl
import re
from nltk.corpus import stopwords

EXCEL_IN = 'forum.no.tags.xlsx'
EXCEL_OUT = 'forum.filtered.xlsx'


def first_word(text):
    if not text:
        return False
    match = re.split(r'\s+', text)
    count = 0
    stop_words = stopwords.words('english')
    for word in match:
        if len(word) > 2 and (word not in stop_words):
            count += 1
            if count > 5:
                # this my be 'i' or 'my' or something else
                return match[0]
    return False


def copy():
    for i in range(1, 10):
        v = posts.cell(sourceRow, i).value
        posts.cell(targetRow, i, v)
        posts.cell(sourceRow, i, '')


def copy_to_personal(from_row, to_row):
    for i in range(1, 10):
        v = posts.cell(from_row, i).value
        xl['Personal'].cell(to_row, i, v)


def clear(src):
    for i in range(1, 10):
        posts.cell(src, i, '')


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']
    targetRow = 2
    sourceRow = 3
    personalRow = 2
    bottomRow = len(posts['G']) + 1
    xl.create_sheet('Personal')
    copy_to_personal(1, 1)  # copy headers
    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value # G column
        fw = first_word(text)
        if fw:
            copy()
            if fw.lower() in ['i', 'my']:
                copy_to_personal(targetRow, personalRow)
                personalRow += 1
            targetRow += 1
        else:
            clear(sourceRow)
            print('"{}"'.format(text))
        sourceRow += 1
    xl.save(EXCEL_OUT)
    xl.close()
