import openpyxl
import re
from nltk.corpus import stopwords

EXCEL_IN = 'forum.filtered.xlsx'
EXCEL_OUT = 'forum.filtered.xlsx'

stop_words = stopwords.words('English')
common = ['scary', 'sleep', 'my head', 'things', 'prolactin', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane',
          'dog bark', 'social', 'delusions', 'sad', 'television', 'tv', 'hallucinations', 'time', 'meds', 'symptoms',
          'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis',
          'headaches','pace', 'pacing', 'voice', 'delusional', 'therapist', 'therapy', 'diagnosed', 'pd', 'mri', 'chronic', 'zyprexa', 'schizophrenia', 'voices', 'sz']  # should be lowercase
personal = ['i', 'my', "i've", "i'm", 'im', 'ive']  # should be lowercase


def filter_object(text):
    obj = {
        'common': False,
        # 'personal': False,
        'long_count': False
    }
    if not text:
        return obj
    match = re.split(r'\s+', text)
    # if match[0].lower() in personal:
    #     obj['personal'] = True
    count = 0
    for word in match:

        if word.lower() in common or ("mg" in word): #lower the word to check against
            obj['common'] = True
        if len(word) > 2 and (word not in stop_words):
            count += 1
    if count > 4:
        obj['long_count'] = True
    return obj


def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10):
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)



if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Non and General']
    sourceRow = 3
    personalRow = 2
    commonRow = 2
    nonAndGenRow = 2
    gmhRow = 2
    nmhRow = 2
    bottomRow = len(posts['G']) + 1
    # xl.create_sheet('Personal')
    # copy_to(1, 1, 'Personal')  # copy headers
    # xl.create_sheet('Personal & Common')
    # copy_to(1, 1, 'Personal & Common')  # copy headers
    # xl.create_sheet('Non and General')
    # copy_to(1, 1, 'Non and General')  # copy headers
    # xl.create_sheet('General Mental Health')
    # copy_to(1, 1, 'General Mental Health')  # copy headers
    # xl.create_sheet('Non Mental Health')
    # copy_to(1, 1, 'Non Mental Health')  # copy headers
    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value # G column
        fo = filter_object(text)
        # if fo['personal'] and fo['long_count']:
        #     copy_to(sourceRow, personalRow, 'Personal')
        #     personalRow += 1
        # if fo['personal'] and fo['long_count'] and fo['common']:
        #     copy_to(sourceRow, commonRow, 'Personal & Common')
        #     commonRow += 1
        # else:
        #     copy_to(sourceRow, nonAndGenRow, 'Non and General')
        #     nonAndGenRow += 1
        if fo['long_count'] and fo['common']:
            copy_to(sourceRow, gmhRow, 'General Mental Health')
            gmhRow += 1
            print("General Mental Health - sectioned")
        else:
            copy_to(sourceRow, nmhRow, 'Non Mental Health')
            nmhRow += 1
            print("Non- Mental Health - sectioned")
        sourceRow += 1
    xl.save(EXCEL_OUT)
    xl.close()
