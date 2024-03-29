import openpyxl
import re

from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords

from nltk.probability import FreqDist
import nltk

EXCEL_IN = 'sample.xlsx'
# EXCEL_IN = 'forum.filtered.xlsx'
#EXCEL_OUT = 'forum.filtered.xlsx'
slist = []
tlist=[]
flist =[]
legal_list =[]

stop_words = stopwords.words("english")

def filter_words(text):
    # if not text:
    #     return False
    match = re.split(r'\s+', text)
    return match





if __name__ == '__main__':    # Script starts here
    new_list = []
    wb = openpyxl.load_workbook('sample.xlsx')
    sourceRow = 3
    rootSourceRow = 3
    personalRow = 3
    commonRow = 3
    nonAndGenRow = 3
    gmhRow = 3
    nmhRow = 3

    posts = wb['Posts']
    bottomRow = 40000

    while sourceRow < bottomRow:
            text = posts.cell(sourceRow, 7).value # G column from Cleaned_Posts Page
            if(text == None):
                pass
            else:
    #
    #
    # wb = openpyxl.load_workbook('sample.xlsx')
    # ws = wb.get_sheet_by_name('Posts')
    # mylist = []
    # for row in ws.iter_rows('G{}:G{}'.format(ws.min_row, ws.max_row)):
    #     for cell in row:
    #         if (cell.value == None):
    #             pass
    #         else:
            # print ("Type: {}, Value: {}".format(type(cell.value), cell.value))
                tokenize = nltk.word_tokenize(text)
                    # print(tokenize)
                # for sent in tokenize:
                #     tokenize_words = re.split(r'\s+', sent)
                    # tokenize_words = nltk.word_tokenize((sent))
                for w in tokenize:
                    if w not in stop_words:
                        new_list.append(w)
            sourceRow += 1

# print(new_list)
print("============================================SPLIT===============================================")

fdist = FreqDist(new_list)
#print(fdist.most_common(200))
for l in (fdist.most_common(5000)):
    print(l)
