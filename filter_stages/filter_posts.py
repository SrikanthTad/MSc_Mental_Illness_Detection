import openpyxl
import re
from nltk.corpus import stopwords

EXCEL_IN = 'forum.no.tags.xlsx'
EXCEL_OUT = 'forum.filtered.xlsx'


def many_words(text):
    if not text:
        return False
    match = re.split(r'\s+', text)
    count = 0
    stop_words = stopwords.words('english')
    stop_words2 = ['I', 'My']
    for word in match:
        if len(word) > 2 and (word not in stop_words):
            count += 1
            if count > 5:
                return True
    return False


def copy(src, tgt):
    for i in range(1, 10):
        v = posts.cell(src, i).value
        posts.cell(tgt, i, v)
        posts.cell(src, i, '')


def clear(src):
    for i in range(1, 10):
        posts.cell(src, i, '')


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']
    targetRow = 2
    sourceRow = 3
    bottomRow = len(posts['G']) + 1
    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value # G column
        if many_words(text):
            copy(sourceRow, targetRow)
            targetRow += 1
        else:
            clear(sourceRow)
            print('"{}"'.format(text))
        sourceRow += 1
        # print([currentRow, lastRow])
    xl.save(EXCEL_OUT)
    xl.close()
