import math
from textblob import TextBlob as tb

def tf(word, blob):
    return blob.words.count(word) / len(blob.words)

def n_containing(word, bloblist):
    return sum(1 for blob in bloblist if word in blob.words)

def idf(word, bloblist):
    return math.log(len(bloblist) / (1 + n_containing(word, bloblist)))

def tfidf(word, blob, bloblist):
    return tf(word, blob) * idf(word, bloblist)


import openpyxl
import re

from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords

from nltk.probability import FreqDist
import nltk

EXCEL_IN = 'forum.filtered.xlsx'
#EXCEL_OUT = 'forum.filtered.xlsx'
slist = []
tlist=[]
flist =[]
legal_list =[]

stop_words = stopwords.words("english")






if __name__ == '__main__':    # Script starts here

    new_list = []
    wb = openpyxl.load_workbook('sample.xlsx')
    sourceRow = 3
    rootSourceRow = 3
    personalRow = 3
    commonRow = 3
    nonAndGenRow = 3
    gmhRow = 3
    nmhRow = 3

    posts = wb['Posts']
    bottomRow = 40000

    while sourceRow < bottomRow:
            text = posts.cell(sourceRow, 7).value # G column from Cleaned_Posts Page
            if(text == None):
                pass
            else:
                # print(text)


                document = tb(str(text))
                    # print(tokenize)
                new_list.append(document)
            sourceRow += 1


    # # wb = openpyxl.load_workbook('forum.filtered.xlsx')
    # wb = openpyxl.load_workbook('sample.xlsx')
    # ws = wb.get_sheet_by_name('Posts')
    # mylist = []
    # for row in ws.iter_rows('G{}:G{}'.format(ws.min_row, ws.max_row)):
    #     for cell in row:
    #         if (cell.value == None):
    #             pass
    #         else:
    #         # print ("Type: {}, Value: {}".format(type(cell.value), cell.value))
    #             document = tb(str(cell.value))
    #             legal_list.append(document)
    #
    #         #     for word in (cell.value.split()):
    #         #         word = word.lower()
    #         #         if word not in stop_words:
    #         #             legal_list.append(word)

print("============================================SPLIT===============================================")
bloblist = new_list


for i, blob in enumerate(bloblist):
    print("Top words in document {}".format(i + 1))
    scores = {word: tfidf(word, blob, bloblist) for word in blob.words}
    sorted_words = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    for word, score in sorted_words[:3]:
        print("\tWord: {}, TF-IDF: {}".format(word, round(score, 5)))



