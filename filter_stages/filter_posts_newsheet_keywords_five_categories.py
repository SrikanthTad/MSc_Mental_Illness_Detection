import openpyxl
import re
from nltk.corpus import stopwords

EXCEL_IN = 'forum.no.tags.xlsx'
EXCEL_OUT = 'forum.filtered_no_sent_length.xlsx'


def filter_words(text):
    if not text:
        return False
    match = re.split(r'\s+', text)
    common = ['scary', 'sleep', 'my head', 'things', 'night', 'echoes', 'antipsychotic', 'whispering', 'alone', 'insane', 'dog bark', 'social', 'delusions', 'sad','television', 'TV', 'hallucinations', 'time', 'meds', 'symptoms', 'voices', 'medication', 'cognitive', 'running', 'thoughts', 'mg', 'scared', 'psychosis', 'headaches'] # should be lowercase
    is_common = False
    for word in match:
        if word in common:
            is_common = True
            break
    personal = ['i', 'my', "i've", "i'm", 'im', 'ive'] # should be lowercase
    return [match[0].lower() in personal, is_common]

def filter_initial(text):
    stop_words = stopwords.words('English')
    if not text:
        return False
    match = re.split(r'\s+', text)
    for word in match:
        if len(word)>2 and (word not in stop_words):
            count = count +1
    if count < 5:
        return False
    else:
        return True


def copy():
    for i in range(1, 10):
        v = posts.cell(sourceRow, i).value
        posts.cell(targetRow, i, v)
        posts.cell(sourceRow, i, '')


def copy_to(from_row, to_row, sheet_name):
    for i in range(1, 10):
        v = posts.cell(from_row, i).value
        xl[sheet_name].cell(to_row, i, v)


def clear(src):
    for i in range(1, 10):
        posts.cell(src, i, '')


if __name__ == '__main__':    # Script starts here
    xl = openpyxl.load_workbook(EXCEL_IN)
    posts = xl['Posts']
    targetRow = 2
    sourceRow = 3
    personalRow = 2
    commonRow = 2
    nonAndGenRow = 2
    bottomRow = len(posts['G']) + 1
    xl.create_sheet('Personal')
    copy_to(1, 1, 'Personal')  # copy headers
    xl.create_sheet('Personal & Common')
    copy_to(1, 1, 'Personal & Common')  # copy headers
    xl.create_sheet('Non and General')
    copy_to(1, 1, 'Non and General')  # copy headers
    while sourceRow < bottomRow:
        text = posts.cell(sourceRow, 7).value # G column
        fw = filter_words(text)
        fy = filter_initial(text)
        if fw:
            copy()
            if fw[0] and fy:
                copy_to(targetRow, personalRow, 'Personal')
                personalRow += 1
                if fw[1] and fy:
                    copy_to(targetRow, commonRow, 'Personal & Common')
                    commonRow += 1
                else:
                    copy_to(targetRow, nonAndGenRow, 'Non and General')
                    nonAndGenRow += 1
            else:
                copy_to(targetRow, nonAndGenRow, 'Non and General')
                nonAndGenRow += 1
            targetRow += 1
        else:
            if text:
                copy_to(sourceRow, nonAndGenRow, 'Non and General')
                nonAndGenRow += 1
            clear(sourceRow)
            print('"{}"'.format(text))
        sourceRow += 1
    xl.save(EXCEL_OUT)
    xl.close()
